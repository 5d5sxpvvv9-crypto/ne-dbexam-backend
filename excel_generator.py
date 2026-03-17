"""
엑셀 생성 모듈 (v3)
10열 고정 포맷:
  문제번호 | 학교 | 학년 | 출제문항 | 공통지문 | 문제지문 | 보기/조건 | 정답 | 문제유형 | 문항형태

공통지문 셀은 해당 문항 범위에서 병합 (MergedCellRange)
"""

import os
import logging
from collections import OrderedDict
from typing import List, Dict, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from question_extractor import QuestionData

logger = logging.getLogger(__name__)

# 10열 고정 컬럼 정의
COLUMNS = [
    {"key": "question_number",  "header": "문제번호", "width": 10},
    {"key": "school",           "header": "학교",     "width": 12},
    {"key": "grade",            "header": "학년",     "width": 8},
    {"key": "question_text",    "header": "출제문항", "width": 45},
    {"key": "common_passage",   "header": "공통지문", "width": 55},
    {"key": "question_passage", "header": "문제지문", "width": 55},
    {"key": "choices",          "header": "보기/조건","width": 45},
    {"key": "answer",           "header": "정답",     "width": 15},
    {"key": "question_type",    "header": "문제유형", "width": 20},
    {"key": "question_format",  "header": "문항형태", "width": 12},
]


def _estimate_line_count(text: str, col_width: int) -> int:
    """wrap_text 적용 시 셀에 표시될 줄 수 추정 (CJK 문자 폭 2배 반영)"""
    if not text:
        return 1
    total = 0
    for segment in text.split('\n'):
        if not segment:
            total += 1
            continue
        display_w = 0
        for ch in segment:
            if '\u3000' <= ch <= '\u9fff' or '\uac00' <= ch <= '\ud7af':
                display_w += 2
            else:
                display_w += 1
        chars_per_line = max(int(col_width * 1.7), 8)
        total += max(1, (display_w + chars_per_line - 1) // chars_per_line)
    return total


def generate_excel(
    all_questions: List[QuestionData],
    output_path: str,
    sheet_name: str = "문제",
    merge_common_passages: bool = True,
) -> str:
    """구조화된 문항 데이터로 엑셀 파일 생성 (9열 고정)"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ── 스타일 ──
    header_font = Font(name="맑은 고딕", bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    data_font = Font(name="맑은 고딕", size=10)
    data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── 헤더 행 ──
    for col_idx, col_def in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_def["header"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

    # ── 시험지(학교·학년) 단위로 그룹핑 ──
    file_groups: OrderedDict[Tuple[str, int], List[QuestionData]] = OrderedDict()
    non_numeric_questions: List[QuestionData] = []

    for q in all_questions:
        try:
            int(q.question_number)
        except (ValueError, TypeError):
            logger.warning(f"문항 번호를 int로 변환 불가: {q.question_number!r} — 말미에 추가")
            non_numeric_questions.append(q)
            continue
        key = (q.school or "", q.grade or 0)
        if key not in file_groups:
            file_groups[key] = []
        file_groups[key].append(q)

    # ── 누락 행 스타일 ──
    missing_font = Font(name="맑은 고딕", size=10, color="999999", italic=True)

    CENTER_KEYS = {"question_number", "school", "grade", "answer", "question_type", "question_format"}

    def _write_row(ws, row_idx, q, is_placeholder=False):
        values = _question_to_row(q)
        for col_idx, col_def in enumerate(COLUMNS, 1):
            value = values.get(col_def["key"], "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = missing_font if is_placeholder else data_font
            cell.border = thin_border
            cell.alignment = center_alignment if col_def["key"] in CENTER_KEYS else data_alignment

    # ── 데이터 행: 각 시험지 그룹별 슬롯 채우기 후 순차 출력 ──
    row_idx = 2
    passage_groups: Dict[int, List[int]] = {}

    for file_key, group_questions in file_groups.items():
        questions_by_no: Dict[int, QuestionData] = {}
        for q in group_questions:
            q_no = int(q.question_number)
            questions_by_no[q_no] = q

        max_no = max(questions_by_no.keys()) if questions_by_no else 0

        for no in range(1, max_no + 1):
            q = questions_by_no.get(no)
            is_placeholder = q is None

            if is_placeholder:
                q = QuestionData(
                    question_number=no,
                    question_text="[MISSING] 문항 누락",
                    question_type="missing",
                    school=file_key[0],
                    grade=file_key[1],
                )
                logger.info(f"누락 문항 placeholder 생성: {file_key[0]} {no}번")

            _write_row(ws, row_idx, q, is_placeholder)

            if not is_placeholder and q.passage_group_id is not None:
                if q.passage_group_id not in passage_groups:
                    passage_groups[q.passage_group_id] = []
                passage_groups[q.passage_group_id].append(row_idx)

            row_idx += 1

    # ── int 변환 불가 문항은 말미에 추가 ──
    for q in non_numeric_questions:
        _write_row(ws, row_idx, q)
        if q.passage_group_id is not None:
            if q.passage_group_id not in passage_groups:
                passage_groups[q.passage_group_id] = []
            passage_groups[q.passage_group_id].append(row_idx)
        row_idx += 1

    # ── 공통지문 셀 병합 (5번째 열 = E열) ──
    if merge_common_passages:
        common_col = 5  # E열 (공통지문)
        for group_id, rows in passage_groups.items():
            if len(rows) >= 2:
                start_row = min(rows)
                end_row = max(rows)
                first_cell = ws.cell(row=start_row, column=common_col)
                if first_cell.value:
                    try:
                        ws.merge_cells(
                            start_row=start_row, start_column=common_col,
                            end_row=end_row, end_column=common_col,
                        )
                        first_cell.alignment = Alignment(
                            horizontal="left", vertical="top", wrap_text=True
                        )
                        logger.info(f"공통지문 셀 병합: E{start_row}:E{end_row} (그룹 {group_id})")
                    except Exception as e:
                        logger.warning(f"셀 병합 실패 (E{start_row}:E{end_row}): {e}")

    # ── 행 높이 자동 조정 (텍스트 wrap 고려) ──
    for r in range(2, row_idx):
        max_lines = 1
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value and isinstance(cell.value, str):
                col_w = COLUMNS[c - 1]["width"]
                max_lines = max(max_lines, _estimate_line_count(cell.value, col_w))
        ws.row_dimensions[r].height = max(15, min(max_lines * 15, 409))

    # ── 틀 고정 ──
    ws.freeze_panes = "A2"

    # ── 저장 ──
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    logger.info(f"엑셀 파일 생성 완료: {output_path} ({row_idx - 2}행)")
    return output_path


def _question_to_row(q: QuestionData) -> dict:
    """QuestionData를 엑셀 행 값(9열)으로 변환"""
    school_name = q.school
    if school_name.endswith("학교"):
        school_name = school_name[:-2]

    return {
        "question_number": q.question_number,
        "school": school_name,
        "grade": q.grade if q.grade else "",
        "question_text": q.question_text,
        "common_passage": q.common_passage,
        "question_passage": q.question_passage,
        "choices": q.choices,
        "answer": q.answer,
        "question_type": q.question_type,
        "question_format": q.question_format,
    }


def generate_merged_excel(
    file_results: List[Dict],
    output_path: str,
) -> str:
    """여러 파일의 결과를 단일 엑셀로 병합"""
    all_questions = []
    for fr in file_results:
        all_questions.extend(fr["questions"])
    return generate_excel(all_questions, output_path)
